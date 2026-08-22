"""Importazione dei dati della lega da CSV.

Il draft si fa di persona e poi si carica il risultato: questo modulo legge il
CSV, lo valida riga per riga e dice esattamente cosa non va e dove, invece di
fallire alla prima cella storta.

Nessuna scrittura avviene se restano errori bloccanti: prima si guarda
l'anteprima, poi si conferma.
"""

from __future__ import annotations

import csv
import difflib
import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date
from typing import TYPE_CHECKING

import pandas as pd

from .regole import RUOLI_MANTRA, ParametriLega

if TYPE_CHECKING:  # pragma: no cover - solo per i type checker
    from .conformita import StatoRosa
    from .modelli import Rosa

# Sinonimi accettati per ogni colonna: i CSV veri arrivano da fogli diversi e
# nessuno scrive le intestazioni allo stesso modo.
COLONNE_ROSE = {
    "squadra": ("squadra", "fantasquadra", "team", "squadrafantacalcio"),
    "giocatore": ("giocatore", "nome", "calciatore", "player"),
    "club": ("club", "squadraseriea", "squadrareale", "seriea"),
    "ruoli": ("ruoli", "ruolo", "rm", "ruolomantra", "mantra"),
    "ingaggio": ("ingaggio", "stipendio", "salario", "salary", "capology"),
    # "costo": nell'export di Leghe Fantacalcio la lega ci scrive gli anni
    # residui, non un prezzo. Convenzione loro, non del sito d'origine.
    "anni": (
        "anni",
        "anniresidui",
        "annicontratto",
        "durata",
        "contratto",
        "costo",
    ),
    "nazionalita": ("nazionalita", "nazione", "paese", "country"),
    "data_nascita": ("datanascita", "nascita", "datadinascita", "compleanno"),
    "prolungato": ("prolungato", "giaprolungato", "prolungamento"),
}

COLONNE_RISULTATI = {
    "giornata": ("giornata", "giornatan", "gg", "round"),
    "casa": ("casa", "squadracasa", "hometeam", "team1"),
    "trasferta": ("trasferta", "squadratrasferta", "ospite", "awayteam", "team2"),
    "punti_casa": ("punticasa", "punteggiocasa", "fantapunticasa"),
    "punti_trasferta": ("puntitrasferta", "punteggiotrasferta", "fantapuntitrasferta"),
}

OBBLIGATORIE_ROSE = ("squadra", "giocatore", "ruoli", "ingaggio", "anni")

# Chiave sotto cui csv.DictReader raccoglie i campi in eccesso di una riga.
EXTRA = "__oltre_le_colonne__"
OBBLIGATORIE_RISULTATI = (
    "giornata",
    "casa",
    "trasferta",
    "punti_casa",
    "punti_trasferta",
)

# Cortesia per chi arriva dal fantacalcio classico: i ruoli si traducono.
RUOLI_CLASSICI = {"P": "Por", "D": "Dc", "C": "C", "A": "A"}


@dataclass(frozen=True)
class Problema:
    """Un errore o un avviso su una riga del CSV."""

    riga: int
    colonna: str
    valore: str
    messaggio: str
    bloccante: bool = True


@dataclass
class EsitoImportazione:
    """Cosa e' stato letto e cosa non torna."""

    righe: list[dict] = field(default_factory=list)
    problemi: list[Problema] = field(default_factory=list)
    intestazioni_ignorate: list[str] = field(default_factory=list)

    @property
    def errori(self) -> list[Problema]:
        return [p for p in self.problemi if p.bloccante]

    @property
    def avvisi(self) -> list[Problema]:
        return [p for p in self.problemi if not p.bloccante]

    @property
    def importabile(self) -> bool:
        return bool(self.righe) and not self.errori


# ---------------------------------------------------------------------------
# Normalizzazione dei valori
# ---------------------------------------------------------------------------


def normalizza_intestazione(testo: str) -> str:
    """'Anni Residui ' -> 'anniresidui': confronta le intestazioni senza sorprese."""
    senza_accenti = "".join(
        c
        for c in unicodedata.normalize("NFD", str(testo))
        if unicodedata.category(c) != "Mn"
    )
    return re.sub(r"[^a-z0-9]", "", senza_accenti.lower())


def mappa_colonne(
    intestazioni: list[str], schema: dict[str, tuple[str, ...]]
) -> tuple[dict[str, str], list[str]]:
    """Associa le intestazioni del CSV ai campi attesi.

    Restituisce (campo -> intestazione originale, intestazioni non riconosciute).
    """
    trovate: dict[str, str] = {}
    ignorate: list[str] = []

    for intestazione in intestazioni:
        chiave = normalizza_intestazione(intestazione)
        for campo, sinonimi in schema.items():
            if chiave in sinonimi and campo not in trovate:
                trovate[campo] = intestazione
                break
        else:
            ignorate.append(intestazione)

    return trovate, ignorate


def leggi_numero(valore: str) -> float:
    """Legge un importo scritto all'italiana.

    Gestisce '3.500.000', '3,5', '€ 2.100.000', '3,5M', '4 mln'. Il suffisso
    milioni e' comune nei fogli di Capology copiati a mano.
    """
    if valore is None:
        raise ValueError("valore mancante")

    testo = str(valore).strip().lower()
    testo = testo.replace("€", "").replace("eur", "").replace(" ", "")
    if not testo:
        raise ValueError("valore mancante")

    moltiplicatore = 1.0
    for suffisso in ("mln", "mil", "m"):
        if testo.endswith(suffisso):
            testo = testo[: -len(suffisso)]
            moltiplicatore = 1_000_000.0
            break

    if moltiplicatore > 1:
        # Con il suffisso milioni sia la virgola sia il punto sono decimali.
        testo = testo.replace(",", ".")
    elif "," in testo:
        # Virgola decimale all'italiana: i punti sono separatori di migliaia.
        testo = testo.replace(".", "").replace(",", ".")
    else:
        # Solo punti: sono migliaia se raggruppano tre cifre (3.500.000).
        if re.fullmatch(r"\d{1,3}(\.\d{3})+", testo):
            testo = testo.replace(".", "")

    try:
        return float(testo) * moltiplicatore
    except ValueError:
        raise ValueError(f"'{valore}' non e' un numero") from None


def leggi_ruoli(valore: str) -> tuple[str, ...]:
    """Accetta 'Dc;Dd', 'Dc/Dd', 'Dc, Dd' e traduce i ruoli classici."""
    if not valore or not str(valore).strip():
        raise ValueError("nessun ruolo indicato")

    grezzi = [r.strip() for r in re.split(r"[;/,|]", str(valore)) if r.strip()]
    ruoli: list[str] = []
    for grezzo in grezzi:
        # Confronto senza badare alle maiuscole: 'dc', 'DC' e 'Dc' sono lo stesso.
        corrispondenza = next(
            (r for r in RUOLI_MANTRA if r.lower() == grezzo.lower()), None
        )
        if corrispondenza is None:
            corrispondenza = RUOLI_CLASSICI.get(grezzo.upper())
        if corrispondenza is None:
            raise ValueError(
                f"ruolo '{grezzo}' sconosciuto (attesi: {', '.join(RUOLI_MANTRA)})"
            )
        if corrispondenza not in ruoli:
            ruoli.append(corrispondenza)

    return tuple(ruoli)


def leggi_data(valore: str) -> date | None:
    """Accetta 2001-04-17, 17/04/2001 e 17-04-2001. Vuoto significa sconosciuta."""
    if valore is None or not str(valore).strip():
        return None

    testo = str(valore).strip()[:10]
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            from datetime import datetime

            return datetime.strptime(testo, formato).date()
        except ValueError:
            continue
    raise ValueError(f"'{valore}' non e' una data riconoscibile (usa gg/mm/aaaa)")


def leggi_booleano(valore: str) -> bool:
    return str(valore).strip().lower() in ("si", "sì", "s", "true", "1", "x", "vero")


# ---------------------------------------------------------------------------
# Lettura delle rose (l'esito del draft)
# ---------------------------------------------------------------------------


def _righe_csv(contenuto: str | bytes) -> tuple[list[str], list[dict]]:
    if isinstance(contenuto, bytes):
        # I fogli esportati da Excel in Italia arrivano spesso in latin-1.
        for codifica in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                contenuto = contenuto.decode(codifica)
                break
            except UnicodeDecodeError:
                continue

    testo = str(contenuto).lstrip("﻿")
    prima_riga = testo.splitlines()[0] if testo.splitlines() else ""
    try:
        dialetto = csv.Sniffer().sniff(prima_riga, delimiters=",;\t")
        separatore = dialetto.delimiter
    except csv.Error:
        separatore = ";" if prima_riga.count(";") > prima_riga.count(",") else ","

    # `restkey=EXTRA` raccoglie i valori in eccesso: succede quando una cella
    # contiene il separatore (tipico: i ruoli scritti "M;C" in un CSV con ';').
    lettore = csv.DictReader(io.StringIO(testo), delimiter=separatore, restkey=EXTRA)
    return list(lettore.fieldnames or []), list(lettore)


def _riga_disallineata(grezza: dict, separatore: str = ";") -> str | None:
    """Messaggio d'errore se la riga ha piu' campi dell'intestazione."""
    extra = grezza.get(EXTRA)
    if not extra:
        return None
    quante = len(extra)
    colonne = "colonna" if quante == 1 else "colonne"
    return (
        f"La riga ha {quante} {colonne} in piu' dell'intestazione: quasi sempre "
        f"significa che una cella contiene il separatore del CSV. Separa i ruoli "
        f"multipli con '/' (per esempio Dc/Dd), non con ';'."
    )


def importa_rose(
    contenuto: str | bytes,
    parametri: ParametriLega | None = None,
    catalogo: dict[str, dict] | None = None,
) -> EsitoImportazione:
    """Legge il CSV delle rose uscito dal draft e valida ogni riga.

    Con un `catalogo` (dal listone ufficiale) il file puo' limitarsi a squadra,
    giocatore, anni e ingaggio: ruoli e club si ricavano dal nome. Un nome che
    non corrisponde a nessun giocatore diventa un errore con dei suggerimenti,
    invece di entrare in silenzio con dati incompleti.
    """
    parametri = parametri or ParametriLega()
    esito = EsitoImportazione()

    intestazioni, righe_grezze = _righe_csv(contenuto)
    if not intestazioni:
        esito.problemi.append(
            Problema(0, "file", "", "Il file e' vuoto o non e' un CSV valido")
        )
        return esito

    colonne, ignorate = mappa_colonne(intestazioni, COLONNE_ROSE)
    esito.intestazioni_ignorate = ignorate

    obbligatorie = (
        ("squadra", "giocatore", "ingaggio", "anni") if catalogo else OBBLIGATORIE_ROSE
    )
    mancanti = [c for c in obbligatorie if c not in colonne]
    if mancanti:
        esito.problemi.append(
            Problema(
                0,
                "intestazioni",
                ", ".join(intestazioni),
                f"Colonne obbligatorie mancanti: {', '.join(mancanti)}",
            )
        )
        return esito

    visti: dict[str, int] = {}

    for numero, grezza in enumerate(righe_grezze, start=2):  # riga 1 = intestazioni

        def campo(nome: str, riga_corrente=grezza) -> str:
            if nome not in colonne:
                return ""
            return (riga_corrente.get(colonne[nome]) or "").strip()

        if disallineata := _riga_disallineata(grezza):
            esito.problemi.append(Problema(numero, "riga", "", disallineata))
            continue

        squadra = campo("squadra")
        giocatore = campo("giocatore")
        if not squadra or not giocatore:
            vuota = not any(
                v.strip() for v in grezza.values() if isinstance(v, str) and v
            )
            if vuota:
                continue  # riga vuota in fondo al foglio: si ignora
            esito.problemi.append(
                Problema(
                    numero,
                    "squadra/giocatore",
                    f"{squadra}/{giocatore}",
                    "Squadra e giocatore sono obbligatori",
                )
            )
            continue

        riga: dict = {"squadra": squadra, "giocatore": giocatore, "club": campo("club")}
        valida = True

        # Con il listone caricato, nome e ruoli si ricavano dal catalogo: il
        # file della lega puo' limitarsi a squadra, giocatore, anni e ingaggio.
        scheda = None
        if catalogo:
            scheda = catalogo.get(normalizza_nome_giocatore(giocatore))
            if scheda is None:
                simili = difflib.get_close_matches(
                    normalizza_nome_giocatore(giocatore), catalogo, n=3, cutoff=0.7
                )
                suggerimento = (
                    " Forse intendevi: " + ", ".join(catalogo[s]["nome"] for s in simili)
                    if simili
                    else ""
                )
                esito.problemi.append(
                    Problema(
                        numero,
                        "giocatore",
                        giocatore,
                        f"Non e' nel listone ufficiale.{suggerimento}",
                    )
                )
                valida = False
            else:
                # Si adotta la grafia del listone: cosi' i nomi restano allineati.
                riga["giocatore"] = scheda["nome"]
                riga["club"] = riga["club"] or scheda["club"]

        if campo("ruoli"):
            try:
                riga["ruoli"] = leggi_ruoli(campo("ruoli"))
            except ValueError as errore:
                esito.problemi.append(
                    Problema(numero, "ruoli", campo("ruoli"), str(errore))
                )
                valida = False
        elif scheda is not None:
            riga["ruoli"] = scheda["ruoli"]
        elif not catalogo:
            # Senza catalogo i ruoli sono obbligatori nel file. Se invece il
            # catalogo c'e' ma il nome non e' stato trovato, l'errore e' gia'
            # stato segnalato sopra: non serve ripeterlo.
            esito.problemi.append(Problema(numero, "ruoli", "", "nessun ruolo indicato"))
            valida = False

        try:
            ingaggio = leggi_numero(campo("ingaggio"))
            if ingaggio < 0:
                raise ValueError("l'ingaggio non puo' essere negativo")
            riga["ingaggio"] = ingaggio
        except ValueError as errore:
            esito.problemi.append(
                Problema(numero, "ingaggio", campo("ingaggio"), str(errore))
            )
            valida = False

        try:
            anni = int(leggi_numero(campo("anni")))
            minimo = parametri.contratto_anni_minimo
            massimo = parametri.contratto_anni_massimo
            if not minimo <= anni <= massimo:
                raise ValueError(
                    f"contratto di {anni} anni: il regolamento ammette da "
                    f"{minimo} a {massimo} (art. 2)"
                )
            riga["anni"] = anni
        except ValueError as errore:
            esito.problemi.append(Problema(numero, "anni", campo("anni"), str(errore)))
            valida = False

        try:
            riga["data_nascita"] = leggi_data(campo("data_nascita"))
        except ValueError as errore:
            esito.problemi.append(
                Problema(numero, "data_nascita", campo("data_nascita"), str(errore))
            )
            valida = False

        riga["nazionalita"] = campo("nazionalita") or "Italia"
        riga["prolungato"] = leggi_booleano(campo("prolungato"))

        if not riga.get("data_nascita"):
            esito.problemi.append(
                Problema(
                    numero,
                    "data_nascita",
                    "",
                    "Senza data di nascita non si puo' verificare lo status "
                    "Under 21 (art. 2)",
                    bloccante=False,
                )
            )
        if not riga["club"]:
            esito.problemi.append(
                Problema(numero, "club", "", "Squadra di Serie A non indicata", False)
            )

        # Articolo 6: un giocatore appartiene a una sola squadra della lega.
        chiave = giocatore.lower()
        if chiave in visti:
            esito.problemi.append(
                Problema(
                    numero,
                    "giocatore",
                    giocatore,
                    f"Gia' presente alla riga {visti[chiave]}: un giocatore puo' "
                    f"stare in una sola rosa",
                )
            )
            valida = False
        else:
            visti[chiave] = numero

        if valida:
            riga["riga_csv"] = numero
            esito.righe.append(riga)

    return esito


def importa_risultati(contenuto: str | bytes) -> EsitoImportazione:
    """Legge il CSV dei risultati di giornata esportato da Leghe Fantacalcio."""
    esito = EsitoImportazione()
    intestazioni, righe_grezze = _righe_csv(contenuto)
    if not intestazioni:
        esito.problemi.append(
            Problema(0, "file", "", "Il file e' vuoto o non e' un CSV valido")
        )
        return esito

    colonne, ignorate = mappa_colonne(intestazioni, COLONNE_RISULTATI)
    esito.intestazioni_ignorate = ignorate

    mancanti = [c for c in OBBLIGATORIE_RISULTATI if c not in colonne]
    if mancanti:
        esito.problemi.append(
            Problema(
                0,
                "intestazioni",
                ", ".join(intestazioni),
                f"Colonne obbligatorie mancanti: {', '.join(mancanti)}",
            )
        )
        return esito

    for numero, grezza in enumerate(righe_grezze, start=2):

        def campo(nome: str, riga_corrente=grezza) -> str:
            return (riga_corrente.get(colonne[nome]) or "").strip()

        if not any(v.strip() for v in grezza.values() if isinstance(v, str) and v):
            continue

        if disallineata := _riga_disallineata(grezza):
            esito.problemi.append(Problema(numero, "riga", "", disallineata))
            continue

        riga: dict = {"casa": campo("casa"), "trasferta": campo("trasferta")}
        valida = bool(riga["casa"] and riga["trasferta"])
        if not valida:
            esito.problemi.append(
                Problema(numero, "casa/trasferta", "", "Servono entrambe le squadre")
            )
            continue

        if riga["casa"].lower() == riga["trasferta"].lower():
            esito.problemi.append(
                Problema(
                    numero,
                    "casa/trasferta",
                    riga["casa"],
                    "Una squadra non puo' giocare contro se stessa",
                )
            )
            continue

        for campo_numerico, etichetta in (
            ("giornata", "giornata"),
            ("punti_casa", "punti casa"),
            ("punti_trasferta", "punti trasferta"),
        ):
            try:
                valore = leggi_numero(campo(campo_numerico))
                riga[campo_numerico] = (
                    int(valore) if campo_numerico == "giornata" else valore
                )
            except ValueError as errore:
                esito.problemi.append(
                    Problema(numero, etichetta, campo(campo_numerico), str(errore))
                )
                valida = False

        if valida and riga["giornata"] < 1:
            esito.problemi.append(
                Problema(numero, "giornata", campo("giornata"), "La giornata parte da 1")
            )
            valida = False

        if valida:
            riga["riga_csv"] = numero
            esito.righe.append(riga)

    # Una squadra puo' giocare una sola partita per giornata: due righe che la
    # citano nello stesso turno sono quasi sempre un copia-incolla sbagliato.
    impegni: dict[tuple[int, str], int] = {}
    for riga in list(esito.righe):
        for nome in (riga["casa"], riga["trasferta"]):
            chiave = (riga["giornata"], nome.strip().lower())
            if chiave in impegni:
                esito.problemi.append(
                    Problema(
                        riga["riga_csv"],
                        "casa/trasferta",
                        nome,
                        f"{nome} compare gia' nella giornata {riga['giornata']} "
                        f"alla riga {impegni[chiave]}",
                    )
                )
                esito.righe.remove(riga)
                break
            impegni[chiave] = riga["riga_csv"]

    return esito


# ---------------------------------------------------------------------------
# Modelli di CSV da scaricare
# ---------------------------------------------------------------------------

MODELLO_ROSE = """squadra;giocatore;club;ruoli;ingaggio;anni;nazionalita;data_nascita
Tiri Team;Rossi Marco;Juventus;Dc;2.500.000;3;Italia;17/04/2001
Tiri Team;Bianchi Luca;Inter;Por;1.800.000;1;Italia;02/11/1996
Padel United;Silva Joao;Napoli;W/T;4,2M;2;Brasile;30/06/1999
"""

MODELLO_RISULTATI = """giornata;casa;trasferta;punti_casa;punti_trasferta
1;Tiri Team;Padel United;72,5;66,0
1;Nuovo Cuneo FC;Real Bisalta;61,0;78,5
"""


def modello_rose() -> str:
    """CSV di esempio con le intestazioni canoniche."""
    return MODELLO_ROSE


def modello_risultati() -> str:
    return MODELLO_RISULTATI


# ---------------------------------------------------------------------------
# Dal CSV alle rose: anteprima e scrittura
# ---------------------------------------------------------------------------


def rose_da_esito(
    esito: EsitoImportazione,
    data_draft: date,
    parametri: ParametriLega | None = None,
) -> dict[str, Rosa]:
    """Costruisce in memoria le rose che il CSV produrrebbe.

    Serve per l'anteprima: si vede se il draft ha lasciato qualcuno fuori dai
    paletti *prima* di scrivere qualsiasi cosa nel database.
    """
    from .modelli import Contratto, Giocatore, Rosa, Squadra

    per_squadra: dict[str, list[dict]] = {}
    for riga in esito.righe:
        per_squadra.setdefault(riga["squadra"], []).append(riga)

    rose: dict[str, Rosa] = {}
    identificativo = 1
    for indice, (nome_squadra, righe) in enumerate(sorted(per_squadra.items()), start=1):
        giocatori: dict[int, Giocatore] = {}
        contratti: list[Contratto] = []
        for riga in righe:
            giocatori[identificativo] = Giocatore(
                id=identificativo,
                nome=riga["giocatore"],
                club=riga.get("club", ""),
                ruoli=riga["ruoli"],
                ingaggio=riga["ingaggio"],
                nazionalita=riga.get("nazionalita", "Italia"),
                data_nascita=riga.get("data_nascita"),
            )
            contratti.append(
                Contratto(
                    giocatore_id=identificativo,
                    squadra_id=indice,
                    anni_residui=riga["anni"],
                    prolungato=riga.get("prolungato", False),
                )
            )
            identificativo += 1

        squadra = Squadra(id=indice, nome=nome_squadra, presidente="")
        rose[nome_squadra] = Rosa(squadra=squadra, contratti=contratti).collega(giocatori)

    return rose


def anteprima_conformita(
    esito: EsitoImportazione,
    data_draft: date,
    parametri: ParametriLega | None = None,
    momento=None,
) -> dict[str, StatoRosa]:
    """Verifica le rose del CSV contro il regolamento, senza scrivere nulla."""
    from .conformita import Momento, verifica_rosa

    momento = momento or Momento.ASTA_SETTEMBRE
    rose = rose_da_esito(esito, data_draft, parametri)
    return {
        nome: verifica_rosa(rosa, data_draft, parametri, momento)
        for nome, rosa in rose.items()
    }


def applica_rose(arch, esito: EsitoImportazione, sostituisci: bool = True) -> dict:
    """Scrive nel database le rose importate.

    Con `sostituisci` le rose precedenti vengono azzerate: e' il caso normale
    dopo un draft, in cui il CSV rappresenta la situazione completa.
    Le squadre gia' esistenti mantengono id e identita' visiva.
    """
    from .data import prossimo_id

    if not esito.importabile:
        raise ValueError(
            f"Importazione non eseguibile: {len(esito.errori)} errori da correggere"
        )

    esistenti = arch.squadre()
    id_per_nome = {}
    if not esistenti.empty:
        id_per_nome = {
            str(r["nome"]).strip().lower(): int(r["id"]) for _, r in esistenti.iterrows()
        }

    nomi_csv = sorted({riga["squadra"] for riga in esito.righe})
    nuove_squadre = []
    prossimo = prossimo_id(arch, "squadre")

    for nome in nomi_csv:
        chiave = nome.strip().lower()
        if chiave not in id_per_nome:
            id_per_nome[chiave] = prossimo
            nuove_squadre.append({"id": prossimo, "nome": nome, "presidente": ""})
            prossimo += 1

    if nuove_squadre:
        arch.scrivi("squadre", nuove_squadre, chiave="id")

    if sostituisci:
        arch.svuota("contratti")
        arch.svuota("giocatori")

    giocatori = []
    contratti = []
    identificativo = prossimo_id(arch, "giocatori")

    for riga in esito.righe:
        squadra_id = id_per_nome[riga["squadra"].strip().lower()]
        nascita = riga.get("data_nascita")
        giocatori.append(
            {
                "id": identificativo,
                "nome": riga["giocatore"],
                "club": riga.get("club", ""),
                "ruoli": ";".join(riga["ruoli"]),
                "ingaggio": riga["ingaggio"],
                "nazionalita": riga.get("nazionalita", "Italia"),
                "data_nascita": nascita.isoformat() if nascita else None,
            }
        )
        contratti.append(
            {
                "giocatore_id": identificativo,
                "squadra_id": squadra_id,
                "anni_residui": riga["anni"],
                "prolungato": int(bool(riga.get("prolungato", False))),
                "stagione_prolungamento": None,
            }
        )
        identificativo += 1

    arch.scrivi("giocatori", giocatori, chiave="id")
    arch.scrivi("contratti", contratti, chiave="giocatore_id")

    return {
        "squadre_create": len(nuove_squadre),
        "giocatori": len(giocatori),
        "contratti": len(contratti),
    }


def applica_risultati(arch, esito: EsitoImportazione) -> dict:
    """Scrive i risultati di giornata, ricavando i gol dalle fasce della lega."""
    from .data import prossimo_id
    from .regole import fasce_gol

    if not esito.importabile:
        raise ValueError(
            f"Importazione non eseguibile: {len(esito.errori)} errori da correggere"
        )

    parametri = ParametriLega()
    squadre = arch.squadre()
    id_per_nome = {
        str(r["nome"]).strip().lower(): int(r["id"]) for _, r in squadre.iterrows()
    }

    sconosciute = {
        nome
        for riga in esito.righe
        for nome in (riga["casa"], riga["trasferta"])
        if nome.strip().lower() not in id_per_nome
    }
    if sconosciute:
        raise ValueError(
            f"Squadre non presenti in lega: {', '.join(sorted(sconosciute))}"
        )

    esistenti = arch.calendario()
    indice_partita = {}
    impegni_esistenti: dict[tuple[int, int], tuple[int, int]] = {}
    if not esistenti.empty:
        for _, r in esistenti.iterrows():
            giornata, casa, trasferta = (
                int(r["giornata"]),
                int(r["casa_id"]),
                int(r["trasferta_id"]),
            )
            indice_partita[(giornata, casa, trasferta)] = int(r["id"])
            impegni_esistenti[(giornata, casa)] = (casa, trasferta)
            impegni_esistenti[(giornata, trasferta)] = (casa, trasferta)
    prossimo = prossimo_id(arch, "calendario") if not esistenti.empty else 1

    righe = []
    for riga in esito.righe:
        casa_id = id_per_nome[riga["casa"].strip().lower()]
        trasferta_id = id_per_nome[riga["trasferta"].strip().lower()]
        chiave = (riga["giornata"], casa_id, trasferta_id)

        if chiave in indice_partita:
            identificativo = indice_partita[chiave]
        else:
            # La partita non esiste con questo accoppiamento: se pero' una delle
            # due squadre e' gia' impegnata in quella giornata, scriverla
            # creerebbe un doppione invece di aggiornare il calendario.
            for squadra_id in (casa_id, trasferta_id):
                gia_impegnata = impegni_esistenti.get((riga["giornata"], squadra_id))
                if gia_impegnata and gia_impegnata != (casa_id, trasferta_id):
                    nome = riga["casa"] if squadra_id == casa_id else riga["trasferta"]
                    raise ValueError(
                        f"Giornata {riga['giornata']}: {nome} risulta gia' impegnata "
                        f"in un'altra partita del calendario. Controlla la giornata "
                        f"o l'ordine casa/trasferta nel CSV."
                    )
            identificativo = prossimo
            prossimo += 1
            impegni_esistenti[(riga["giornata"], casa_id)] = (casa_id, trasferta_id)
            impegni_esistenti[(riga["giornata"], trasferta_id)] = (casa_id, trasferta_id)

        righe.append(
            {
                "id": identificativo,
                "giornata": riga["giornata"],
                "casa_id": casa_id,
                "trasferta_id": trasferta_id,
                "punti_casa": riga["punti_casa"],
                "punti_trasferta": riga["punti_trasferta"],
                "gol_casa": fasce_gol(riga["punti_casa"], parametri),
                "gol_trasferta": fasce_gol(riga["punti_trasferta"], parametri),
            }
        )

    arch.scrivi("calendario", righe, chiave="id")
    return {"partite": len(righe)}


# ---------------------------------------------------------------------------
# Listone ufficiale delle quotazioni (xlsx di Fantacalcio.it)
# ---------------------------------------------------------------------------

# Il file ha una riga di titolo e le intestazioni vere alla seconda riga.
RIGA_INTESTAZIONI_LISTONE = 2
FOGLIO_LISTONE = "Tutti"


def normalizza_nome_giocatore(nome: str) -> str:
    """Chiave di confronto fra il listone e i file della lega.

    'Martinez Jo.' e 'MARTINEZ JO' devono corrispondere: si tolgono accenti,
    punteggiatura e spazi.
    """
    senza_accenti = "".join(
        c
        for c in unicodedata.normalize("NFD", str(nome))
        if unicodedata.category(c) != "Mn"
    )
    return re.sub(r"[^a-z0-9]", "", senza_accenti.lower())


def importa_listone(contenuto: bytes) -> EsitoImportazione:
    """Legge il listone ufficiale: anagrafica, ruoli Mantra e quotazioni.

    Non contiene ne' le assegnazioni alle squadre ne' gli ingaggi: serve a
    popolare il catalogo dei giocatori, non le rose.
    """
    import io as _io

    from openpyxl import load_workbook

    esito = EsitoImportazione()
    try:
        cartella = load_workbook(_io.BytesIO(contenuto), data_only=True, read_only=True)
    except Exception as errore:  # noqa: BLE001 - openpyxl alza tipi eterogenei
        esito.problemi.append(
            Problema(0, "file", "", f"Non e' un file Excel leggibile: {errore}")
        )
        return esito

    foglio = (
        cartella[FOGLIO_LISTONE]
        if FOGLIO_LISTONE in cartella.sheetnames
        else cartella[cartella.sheetnames[0]]
    )

    righe = list(foglio.iter_rows(values_only=True))
    if len(righe) <= RIGA_INTESTAZIONI_LISTONE:
        esito.problemi.append(Problema(0, "file", "", "Il foglio e' vuoto"))
        return esito

    intestazioni = [
        str(v).strip() if v is not None else ""
        for v in righe[RIGA_INTESTAZIONI_LISTONE - 1]
    ]
    indice = {
        normalizza_intestazione(nome): posizione
        for posizione, nome in enumerate(intestazioni)
    }
    esito.intestazioni_ignorate = []

    def colonna(*chiavi: str):
        for chiave in chiavi:
            if chiave in indice:
                return indice[chiave]
        return None

    posizioni = {
        "id_ufficiale": colonna("id"),
        "ruoli": colonna("rm", "ruolomantra"),
        "nome": colonna("nome"),
        "club": colonna("squadra"),
        "quotazione": colonna("qtam", "qta"),
        "fvm": colonna("fvmm", "fvm"),
    }
    mancanti = [
        k for k in ("id_ufficiale", "ruoli", "nome", "club") if posizioni[k] is None
    ]
    if mancanti:
        esito.problemi.append(
            Problema(
                0,
                "intestazioni",
                ", ".join(i for i in intestazioni if i),
                f"Colonne del listone non trovate: {', '.join(mancanti)}. "
                f"Attese Id, RM, Nome, Squadra.",
            )
        )
        return esito

    visti: set[int] = set()
    for numero, grezza in enumerate(
        righe[RIGA_INTESTAZIONI_LISTONE:], start=RIGA_INTESTAZIONI_LISTONE + 1
    ):

        def valore(campo: str, riga=grezza):
            posizione = posizioni[campo]
            if posizione is None or posizione >= len(riga):
                return None
            return riga[posizione]

        if valore("nome") in (None, ""):
            continue

        try:
            id_ufficiale = int(valore("id_ufficiale"))
        except (TypeError, ValueError):
            esito.problemi.append(
                Problema(numero, "Id", str(valore("id_ufficiale")), "Id non numerico")
            )
            continue

        if id_ufficiale in visti:
            continue  # il foglio "Tutti" ripete i giocatori dei fogli per ruolo
        visti.add(id_ufficiale)

        try:
            ruoli = leggi_ruoli(str(valore("ruoli")))
        except ValueError as errore:
            esito.problemi.append(
                Problema(numero, "RM", str(valore("ruoli")), str(errore))
            )
            continue

        def numero_o_none(campo: str) -> float | None:
            grezzo = valore(campo)
            try:
                return float(grezzo) if grezzo not in (None, "") else None
            except (TypeError, ValueError):
                return None

        esito.righe.append(
            {
                "id_ufficiale": id_ufficiale,
                "nome": str(valore("nome")).strip(),
                "club": str(valore("club") or "").strip(),
                "ruoli": ruoli,
                "quotazione": numero_o_none("quotazione"),
                "fvm": numero_o_none("fvm"),
                "riga_csv": numero,
            }
        )

    if not esito.righe:
        esito.problemi.append(
            Problema(0, "file", "", "Nessun giocatore leggibile nel listone")
        )
    return esito


def applica_listone(arch, esito: EsitoImportazione) -> dict:
    """Scrive il catalogo dei giocatori, conservando ingaggi e contratti.

    Un giocatore gia' presente (stesso Id ufficiale) viene aggiornato nei campi
    del listone; ingaggio, nazionalita' e data di nascita restano quelli che
    erano, perche' il listone non li contiene.
    """
    from .data import prossimo_id

    if not esito.importabile:
        raise ValueError(
            f"Importazione non eseguibile: {len(esito.errori)} errori da correggere"
        )

    esistenti = arch.giocatori()
    per_ufficiale: dict[int, dict] = {}
    if not esistenti.empty and "id_ufficiale" in esistenti.columns:
        for _, r in esistenti.iterrows():
            if r["id_ufficiale"] is not None and not pd.isna(r["id_ufficiale"]):
                per_ufficiale[int(r["id_ufficiale"])] = r.to_dict()

    prossimo = prossimo_id(arch, "giocatori")
    righe = []
    aggiornati = 0
    for riga in esito.righe:
        precedente = per_ufficiale.get(riga["id_ufficiale"])
        if precedente is not None:
            identificativo = int(precedente["id"])
            ingaggio = float(precedente.get("ingaggio") or 0.0)
            nazionalita = precedente.get("nazionalita") or "Italia"
            nascita = precedente.get("data_nascita")
            aggiornati += 1
        else:
            identificativo = prossimo
            prossimo += 1
            ingaggio, nazionalita, nascita = 0.0, "Italia", None

        righe.append(
            {
                "id": identificativo,
                "id_ufficiale": riga["id_ufficiale"],
                "nome": riga["nome"],
                "club": riga["club"],
                "ruoli": ";".join(riga["ruoli"]),
                "ingaggio": ingaggio,
                "nazionalita": nazionalita,
                "data_nascita": nascita if isinstance(nascita, str) else None,
                "quotazione": riga["quotazione"],
                "fvm": riga["fvm"],
            }
        )

    arch.scrivi("giocatori", righe, chiave="id")
    return {
        "totali": len(righe),
        "nuovi": len(righe) - aggiornati,
        "aggiornati": aggiornati,
    }


def catalogo_giocatori(arch) -> dict[str, dict]:
    """Giocatori indicizzati per nome normalizzato, per risolvere le rose."""
    righe = arch.giocatori()
    if righe.empty:
        return {}
    catalogo: dict[str, dict] = {}
    for _, r in righe.iterrows():
        catalogo[normalizza_nome_giocatore(r["nome"])] = {
            "nome": r["nome"],
            "club": r["club"],
            "ruoli": tuple(str(r["ruoli"]).split(";")),
            "quotazione": r.get("quotazione"),
        }
    return catalogo
